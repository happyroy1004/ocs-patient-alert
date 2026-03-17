# excel_utils.py

import streamlit as st
import pandas as pd
import io
import msoffcrypto
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from config import PROFESSORS_DICT, SHEET_KEYWORD_TO_DEPARTMENT_MAP

# --- Firebase 연동 함수 ---
def load_all_registered_pids(db_ref_func):
    """
    Firebase에서 모든 사용자가 등록한 환자의 진료번호(PID)와 등록된 진료과 목록을 로드합니다.
    Firebase 구조: {user_key: {PID: {교정: true, ...}, ...}}
    반환 형식: {'PID1': ['교정', '보존'], 'PID2': ['소치'], ...}
    """
    try:
        all_patients_by_user = db_ref_func("patients").get() 
        registered_pids_with_depts = {}
        
        standard_dept_names = set(SHEET_KEYWORD_TO_DEPARTMENT_MAP.values())
        standard_dept_keys = {name.lower() for name in standard_dept_names} # 소문자 진료과 키 셋
        
        if all_patients_by_user:
            # 1. 사용자별 환자 목록 순회 (user_key: 'asteriajimin619_at_gmail_dot_com')
            for user_key, user_patients in all_patients_by_user.items():
                
                if user_patients and isinstance(user_patients, dict):
                    # 2. 환자 진료번호(PID)별 정보 순회 (pid_key: '100203')
                    for pid_key, patient_info in user_patients.items(): 
                        
                        # PID와 patient_info 유효성 검사
                        if not pid_key or not isinstance(pid_key, str) or not isinstance(patient_info, dict):
                            continue
                        
                        pid = pid_key.strip()
                        current_depts = registered_pids_with_depts.get(pid, set())
                        
                        # 3. 진료과 플래그 확인: patient_info의 모든 키를 순회하며 표준 진료과 이름과 매칭
                        for key, value in patient_info.items():
                            # 키를 소문자로 변환하여 표준 진료과 키와 일치하는지 확인
                            key_lower = str(key).lower()
                            
                            # 해당 키가 표준 진료과 키 목록에 포함되고, 값이 True인지 확인
                            if key_lower in standard_dept_keys and value in [True, 'true']:
                                # 표준화된 진료과 이름(예: '교정')을 찾아서 Set에 추가
                                for dept_name in standard_dept_names:
                                    if dept_name.lower() == key_lower:
                                        current_depts.add(dept_name)
                                        break
                                
                        registered_pids_with_depts[pid] = current_depts
            
            # 🚨 디버깅을 위해 추가 (실제 운영 시 주석 처리 권장)
            # st.info(f"🚨 디버그: Firebase에서 총 {len(registered_pids_with_depts)}개의 유니크 PID를 로드했습니다. (예시: {list(registered_pids_with_depts.keys())[:3]})")

        # Set을 List로 변환하여 반환
        return {pid: list(depts) for pid, depts in registered_pids_with_depts.items()}
    except Exception as e:
        # st.error(f"🚨 디버그 오류: Firebase 환자 데이터 로드 중 오류 발생: {e}")
        return {} # 오류 발생 시 빈 딕셔너리 반환

# --- 유효성 검사 ---
def is_daily_schedule(file_name):
    """OCS 스케줄 파일 이름 형식(ocs_YYYY.xlsx/xlsm)을 확인합니다."""
    pattern = r'^ocs_\d{4}\.(?:xlsx|xlsm)$'
    return re.match(pattern, file_name, re.IGNORECASE) is not None
    
def is_encrypted_excel(file_path):
    """엑셀 파일이 암호화되었는지 확인합니다."""
    try:
        file_path.seek(0)
        return msoffcrypto.OfficeFile(file_path).is_encrypted()
    except Exception:
        return False

# --- 엑셀 로드 및 복호화 ---
def load_excel(file, password=None):
    """업로드된 엑셀 파일을 로드하고 필요시 복호화합니다."""
    try:
        file.seek(0)
        file_bytes = file.read()
        
        input_stream = io.BytesIO(file_bytes)
        decrypted_bytes_io = None
        
        # 파일이 암호화되었는지 확인
        is_encrypted = False
        try:
            if msoffcrypto.OfficeFile(input_stream).is_encrypted():
                is_encrypted = True
        except:
            pass
        
        if is_encrypted:
            if not password:
                raise ValueError("암호화된 파일입니다. 비밀번호를 입력해주세요.")
            
            decrypted_bytes_io = io.BytesIO()
            input_stream.seek(0)
            
            office_file = msoffcrypto.OfficeFile(input_stream)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_bytes_io)
            
            decrypted_bytes_io.seek(0)
            return pd.ExcelFile(decrypted_bytes_io), decrypted_bytes_io

        else:
            input_stream.seek(0)
            return pd.ExcelFile(input_stream), input_stream
            
    except Exception as e:
        raise ValueError(f"엑셀 로드 또는 복호화 실패: {e}")

# --- 데이터 처리 및 정렬 ---
def process_sheet_v8(df, professors_list, sheet_key): 
    """OCS 시트 데이터를 교수/비교수 기준으로 정렬합니다."""
    
    required_cols = ['진료번호', '예약일시', '예약시간', '환자명', '예약의사', '진료내역']
    if not all(col in df.columns for col in ['예약의사', '예약시간']):
        return pd.DataFrame(columns=[col for col in required_cols if col in df.columns])

    df = df.sort_values(by=['예약의사', '예약시간'])
    professors = df[df['예약의사'].isin(professors_list)]
    non_professors = df[~df['예약의사'].isin(professors_list)]

    # 정렬 기준 설정
    if sheet_key != '울랄라':
        non_professors = non_professors.sort_values(by=['예약시간', '예약의사'])
    else:
        non_professors = non_professors.sort_values(by=['예약의사', '예약시간'])

    final_rows = []
    current_time = None
    current_doctor = None

    # 비교수 환자 처리 (시간 또는 의사별 그룹핑)
    for _, row in non_professors.iterrows():
        if sheet_key != '울랄라':
            if current_time != row['예약시간']:
                if current_time is not None:
                    final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
                current_time = row['예약시간']
        else:
            if current_doctor != row['예약의사']:
                if current_doctor is not None:
                    final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
                current_doctor = row['예약의사']
        final_rows.append(row)

    # 교수님 섹션 구분자 추가
    if not non_professors.empty:
        final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
    final_rows.append(pd.Series(["<교수님>"] + [" "] * (len(df.columns) - 1), index=df.columns))

    # 교수 환자 처리 (의사별 그룹핑)
    current_professor = None
    for _, row in professors.iterrows():
        if current_professor != row['예약의사']:
            if current_professor is not None:
                final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
            current_professor = row['예약의사']
        final_rows.append(row)

    final_df = pd.DataFrame(final_rows, columns=df.columns)
    final_df = final_df[[col for col in required_cols if col in final_df.columns]]
    return final_df

def process_excel_file_and_style(file_bytes_io, db_ref_func):
    """엑셀 파일을 읽고, 정렬/스타일링을 적용한 후, 분석용 DataFrame 딕셔너리를 반환합니다."""
    file_bytes_io.seek(0)
    output_buffer_for_styling = io.BytesIO()

    try:
        wb_raw = load_workbook(filename=file_bytes_io, keep_vba=False, data_only=True)
    except Exception as e:
        raise ValueError(f"엑셀 워크북 로드 실패: {e}")

    # 1. Firebase에서 등록된 모든 환자 진료번호(PID)와 등록된 진료과 로드
    registered_pids_with_depts = load_all_registered_pids(db_ref_func)
    
    # 2. 회색 스타일 정의
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    processed_sheets_dfs = {}
    cleaned_raw_dfs = {}
    
    # 1. 시트별 데이터 처리 및 정렬
    for sheet_name_raw in wb_raw.sheetnames:
        sheet_name_lower = sheet_name_raw.strip().lower()

        # 시트 이름 매핑
        sheet_key = None
        for keyword, department_name in sorted(SHEET_KEYWORD_TO_DEPARTMENT_MAP.items(), key=lambda item: len(item[0]), reverse=True):
            if keyword.lower() in sheet_name_lower:
                sheet_key = department_name
                break
        if not sheet_key: continue

        ws = wb_raw[sheet_name_raw]
        values = list(ws.values)
        while values and (values[0] is None or all((v is None or str(v).strip() == "") for v in values[0])):
            values.pop(0)
        if len(values) < 2: continue

        df = pd.DataFrame(values)
        if df.empty or df.iloc[0].isnull().all(): continue

        df.columns = df.iloc[0]
        df = df.drop([0]).reset_index(drop=True)
        df = df.fillna("").astype(str)

        if '예약의사' not in df.columns: continue
        df['예약의사'] = df['예약의사'].str.strip().str.replace(" 교수님", "", regex=False)
        
        cleaned_raw_dfs[sheet_name_raw] = df.copy() 

        professors_list = PROFESSORS_DICT.get(sheet_key, [])
        
        try:
            # 정렬된 데이터프레임 생성
            processed_df = process_sheet_v8(df.copy(), professors_list, sheet_key)
            processed_sheets_dfs[sheet_name_raw] = processed_df
        except Exception as e:
            continue

    if not processed_sheets_dfs:
        if cleaned_raw_dfs:
            return cleaned_raw_dfs, None
        file_bytes_io.seek(0)
        all_sheet_dfs = pd.read_excel(file_bytes_io, sheet_name=None)
        return all_sheet_dfs, None

    # 2. 정렬된 데이터로 새 엑셀 파일 생성 및 스타일링
    with pd.ExcelWriter(output_buffer_for_styling, engine='openpyxl') as writer:
        for sheet_name_raw, df in processed_sheets_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name_raw, index=False)

    output_buffer_for_styling.seek(0)
    wb_styled = load_workbook(output_buffer_for_styling, keep_vba=False, data_only=True)

    # 스타일링 로직
    for sheet_name in wb_styled.sheetnames:
        ws = wb_styled[sheet_name]
        
        # 헤더 값을 문자열로 변환하고 공백을 제거하여 안정적인 딕셔너리 생성
        header = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
        
        # 시트 이름에서 현재 진료과(sheet_dept) 추출
        sheet_dept = None
        sheet_name_lower = sheet_name.strip().lower()
        for keyword, department_name in sorted(SHEET_KEYWORD_TO_DEPARTMENT_MAP.items(), key=lambda item: len(item[0]), reverse=True):
            if keyword.lower() in sheet_name_lower:
                sheet_dept = department_name # 표준화된 진료과 이름 (예: '교정', '소치')
                break
        
        # PID 컬럼 인덱스 찾기
        pid_col_idx = None
        for key in ['진료번호', '환자번호', '차트번호', 'PID']:
            if header.get(key):
                pid_col_idx = header.get(key)
                break
        
        # PID 컬럼을 찾지 못했거나 진료과가 매칭되지 않았으면 스킵
        if not pid_col_idx or not sheet_dept:
            continue

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            
            is_registered_patient = False
            
            # 환자 등록 여부에 따른 회색 스타일링
            if pid_col_idx and len(row) >= pid_col_idx:
                 pid_cell = row[pid_col_idx - 1]
                 pid_raw_value = pid_cell.value

                 # 💡 PID 형식 통일 로직 수정 (앞의 0을 제거하고 숫자로만 변환)
                 pid_str = str(pid_raw_value).strip()
                 
                 # .0이 붙은 float 문자열을 int로 변환
                 if pid_str.endswith('.0'):
                    pid_str = pid_str[:-2]
                    
                 # Scientific notation (예: 1.02896E+07) 처리
                 if 'E' in pid_str.upper():
                    try:
                        pid_str = str(int(float(pid_str)))
                    except ValueError:
                        pass # 변환 실패 시 기존 문자열 유지
                
                 # 최종적으로 숫자만 추출하고, 앞의 0을 제거하기 위해 int로 변환 후 다시 문자열로 변환
                 pid_value_digits = "".join(filter(str.isdigit, pid_str))
                 
                 # 🚨 핵심 수정: 정수로 변환 후 다시 문자열로 만들어 앞의 0을 완전히 제거
                 try:
                     pid_value = str(int(pid_value_digits)) 
                 except ValueError:
                     pid_value = pid_value_digits # 숫자가 아닐 경우 기존 값 유지

                 # 매칭 조건 강화: 1. PID가 등록되어 있고, 2. 현재 시트 진료과가 등록된 진료과 목록에 포함되어야 함
                 registered_depts = registered_pids_with_depts.get(pid_value)
                 
                 if (registered_depts and 
                     sheet_dept in registered_depts and 
                     str(row[0].value).strip() not in ["", "<교수님>"]):
                    
                    is_registered_patient = True
                    for cell in row:
                        cell.fill = gray_fill # 회색 배경 적용
                        
            # 교수님 섹션 구분자 스타일링
            if row[0].value == "<교수님>":
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True)

            # 교정 Bonding 강조 스타일
            if sheet_name.strip() == "교정" and '진료내역' in header:
                idx = header['진료내역'] - 1
                if len(row) > idx:
                    cell = row[idx]
                    text = str(cell.value).strip().lower()
                    
                    if ('bonding' in text or '본딩' in text) and 'debonding' not in text:
                        # 회색 배경이 적용되지 않은 경우에만 폰트 스타일 적용
                        if not is_registered_patient: 
                            cell.font = Font(bold=True)

    final_output_bytes = io.BytesIO()
    wb_styled.save(final_output_bytes)
    final_output_bytes.seek(0)
    
    return cleaned_raw_dfs, final_output_bytes

# --- OCS 데이터 분석 ---
def run_analysis(df_dict):
    """OCS 데이터를 기반으로 소치/보존/교정의 통계를 분석합니다."""
    analysis_results = {}
    
    # 분석에 필요한 시트 이름 매핑
    sheet_department_map = {
        '소치': '소치', '소아치과': '소치', '소아 치과': '소치', '보존': '보존', '보존과': '보존', '치과보존과': '보존',
        '교정': '교정', '교정과': '교정', '치과교정과': '교정'
    }
    
    mapped_dfs = {}
    for sheet_name, df in df_dict.items():
        processed_sheet_name = sheet_name.replace(" ", "").lower()
        for key, dept in sheet_department_map.items():
            if processed_sheet_name == key.replace(" ", "").lower():
                if all(col in df.columns for col in ['예약의사', '예약시간', '진료내역']):
                     mapped_dfs[dept] = df.copy()
                break

    # 1. 소치 분석
    if '소치' in mapped_dfs:
        df = mapped_dfs['소치']
        non_professors_df = df[~df['예약의사'].isin(PROFESSORS_DICT.get('소치', []))]
        non_professors_df['예약시간'] = non_professors_df['예약시간'].astype(str).str.strip()
        non_professors_df = non_professors_df[non_professors_df['예약시간'].str.contains(':')] 
        
        # 오전: 08:00 ~ 12:50
        morning_patients = non_professors_df[(non_professors_df['예약시간'] >= '08:00') & (non_professors_df['예약시간'] <= '12:50')].shape[0]
        # 오후: 13:00 이후
        afternoon_patients = non_professors_df[non_professors_df['예약시간'] >= '13:00'].shape[0]
        
        analysis_results['소치'] = {'오전': morning_patients, '오후': afternoon_patients}

    # 2. 보존 분석
    if '보존' in mapped_dfs:
        df = mapped_dfs['보존']
        non_professors_df = df[~df['예약의사'].isin(PROFESSORS_DICT.get('보존', []))]
        non_professors_df['예약시간'] = non_professors_df['예약시간'].astype(str).str.strip()
        non_professors_df = non_professors_df[non_professors_df['예약시간'].str.contains(':')]

        # 오전: 08:00 ~ 12:30
        morning_patients = non_professors_df[(non_professors_df['예약시간'] >= '08:00') & (non_professors_df['예약시간'] <= '12:30')].shape[0]
        # 오후: 12:50 이후
        afternoon_patients = non_professors_df[non_professors_df['예약시간'] >= '12:50'].shape[0]
        
        analysis_results['보존'] = {'오전': morning_patients, '오후': afternoon_patients}

    # 3. 교정 분석 (Bonding)
    if '교정' in mapped_dfs:
        df = mapped_dfs['교정']
        # Bonding이 포함되고 debonding이 없는 환자 필터링
        bonding_patients_df = df[df['진료내역'].str.contains('bonding|본딩', case=False, na=False) & ~df['진료내역'].str.contains('debonding', case=False, na=False)]
        bonding_patients_df['예약시간'] = bonding_patients_df['예약시간'].astype(str).str.strip()
        bonding_patients_df = bonding_patients_df[bonding_patients_df['예약시간'].str.contains(':')]

        # 오전: 08:00 ~ 12:30
        morning_bonding_patients = bonding_patients_df[(bonding_patients_df['예약시간'] >= '08:00') & (bonding_patients_df['예약시간'] <= '12:30')].shape[0]
        # 오후: 12:50 이후
        afternoon_bonding_patients = bonding_patients_df[bonding_patients_df['예약시간'] >= '12:50'].shape[0]
        
        analysis_results['교정'] = {'오전': morning_bonding_patients, '오후': afternoon_bonding_patients}
        
    return analysis_results
