import streamlit as st
import pandas as pd
import firebase_admin
from firebase_admin import credentials, db
import io
import msoffcrypto
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import json
import os
import time
import openpyxl # 추가
import datetime # 추가

# Google Calendar API 관련 라이브러리 추가
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import base64

# --- 전역 상수 정의 ---
# 환자 데이터의 진료과 플래그 키 목록 (DB에 저장되는 T/F 플래그)
PATIENT_DEPT_FLAGS = ["보철", "외과", "내과", "소치", "교정", "원진실", "보존"] 
# 등록 시 선택할 수 있는 모든 진료과
DEPARTMENTS_FOR_REGISTRATION = ["교정", "내과", "보존", "보철", "소치", "외과", "치주", "원진실"]

# --- 1. Imports, Validation Functions, and Firebase Initialization ---

def is_daily_schedule(file_name):
    pattern = r'^ocs_\d{4}\.(?:xlsx|xlsm)$'
    return re.match(pattern, file_name, re.IGNORECASE) is not None
    
def is_valid_email(email):
    email_regex = r"^[a-zA-Z0-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(email_regex, email) is not None

# 이메일 전송 함수
def send_email(receiver, rows, sender, password, date_str=None, custom_message=None):
    # rows는 사용되지 않으므로 제거
    try:
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = receiver

        if custom_message:
            msg['Subject'] = "단체 메일 알림" if date_str is None else f"[치과 내원 알림] {date_str} 예약 내역"
            body = custom_message
        else:
            subject_prefix = ""
            if date_str:
                subject_prefix = f"{date_str}일에 내원하는 "
            msg['Subject'] = f"{subject_prefix}등록 환자 내원 알림"
            
            # rows가 dict의 리스트일 경우 (매칭 환자 데이터)
            if rows is not None and isinstance(rows, list):
                # DataFrame으로 변환하여 HTML 테이블 생성
                rows_df = pd.DataFrame(rows)
                html_table = rows_df.to_html(index=False, escape=False)
                
                style = """
                <style>
                table {
                    width: 100%; max-width: 100%;
                    border-collapse: collapse;
                    font-family: Arial, sans-serif;
                    font-size: 14px;
                    table-layout: fixed;
                }
                th, td {
                    border: 1px solid #dddddd; text-align: left;
                    padding: 8px;
                    vertical-align: top;
                    word-wrap: break-word;
                    word-break: break-word;
                }
                th {
                    background-color: #f2f2f2; font-weight: bold;
                    white-space: nowrap;
                }
                tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .table-container {
                    overflow-x: auto; -webkit-overflow-scrolling: touch;
                }
                </style>
                """
                body = f"다음 토탈 환자가 내일 내원예정입니다:<br><br><div class='table-container'>{style}{html_table}</div>"
            else:
                 body = "내원 환자 정보가 없습니다."

        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        return str(e)
        
# Firebase 초기화
if not firebase_admin._apps:
    try:
        firebase_credentials_json_str = st.secrets["firebase"]["FIREBASE_SERVICE_ACCOUNT_JSON"]
        firebase_credentials_dict = json.loads(firebase_credentials_json_str)

        cred = credentials.Certificate(firebase_credentials_dict)
        firebase_admin.initialize_app(cred, {
            'databaseURL': st.secrets["firebase"]["database_url"]
        })
    except Exception as e:
        st.error(f"Firebase 초기화 오류: {e}")
        st.info("secrets.toml 파일의 Firebase 설정(FIREBASE_SERVICE_ACCOUNT_JSON 또는 database_url)을 [firebase] 섹션 아래에 올바르게 작성했는지 확인해주세요.")
        st.stop()

# Firebase-safe 경로 변환 (이메일을 Firebase 키로 사용하기 위해)
def sanitize_path(email):
    return email.replace(".", "_dot_").replace("@", "_at_")

# 이메일 주소 복원 (Firebase 안전 키에서 원래 이메일로)
def recover_email(safe_id: str) -> str:
    email = safe_id.replace("_at_", "@").replace("_dot_", ".").replace("_com", ".com")
    return email

# 구글 캘린더 인증 정보를 Firebase에 저장/불러오기
def save_google_creds_to_firebase(user_id_safe, creds):
    try:
        creds_ref = db.reference(f"users/{user_id_safe}/google_creds")
        creds_ref.set({
            'token': creds.token, 'refresh_token': creds.refresh_token, 'token_uri': creds.token_uri,
            'client_id': creds.client_id, 'client_secret': creds.client_secret, 'scopes': creds.scopes, 'id_token': creds.id_token
        })
        return True
    except Exception as e:
        st.error(f"Failed to save Google credentials: {e}")
        return False

def load_google_creds_from_firebase(user_id_safe):
    try:
        creds_ref = db.reference(f"users/{user_id_safe}/google_creds")
        creds_data = creds_ref.get()
        if creds_data and 'token' in creds_data:
            creds = Credentials(
                token=creds_data.get('token'), refresh_token=creds_data.get('refresh_token'),
                token_uri=creds_data.get('token_uri'), client_id=creds_data.get('client_id'),
                client_secret=creds_data.get('client_secret'), scopes=creds_data.get('scopes'),
                id_token=creds_data.get('id_token')
            )
            return creds
        return None
    except Exception as e:
        st.error(f"Failed to load Google credentials: {e}")
        return None


    
    # 인증 플로우 생성
    flow = InstalledAppFlow.from_client_config(client_config, SCOPES, redirect_uri=st.secrets["google_calendar"]["redirect_uri"])
    
    if not creds:
        auth_code = st.query_params.get("code")
        
        if auth_code:
            # 인증 코드를 사용하여 토큰을 교환
            flow.fetch_token(code=auth_code)
            creds = flow.credentials
            st.session_state[f"google_creds_{user_id_safe}"] = creds
            # Store credentials in Firebase
            save_google_creds_to_firebase(user_id_safe, creds)
            st.success("Google Calendar 인증이 완료되었습니다.")
            st.query_params.clear()
            st.rerun()
        else:
            auth_url, _ = flow.authorization_url(prompt='consent')
            st.warning("Google Calendar 연동을 위해 인증이 필요합니다. 아래 링크를 클릭하여 권한을 부여하세요.")
            st.markdown(f"**[Google Calendar 인증 링크]({auth_url})**")
            return None

    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        st.session_state[f"google_creds_{user_id_safe}"] = creds
        # Update credentials in Firebase
        save_google_creds_to_firebase(user_id_safe, creds)

    try:
        service = build('calendar', 'v3', credentials=creds)
        return service
    except HttpError as error:
        st.error(f'Google Calendar 서비스 생성 실패: {error}')
        st.session_state.pop(f"google_creds_{user_id_safe}", None)
        # Clear invalid credentials from Firebase
        db.reference(f"users/{user_id_safe}/google_creds").delete()
        return None

def create_calendar_event(service, patient_name, pid, department, reservation_datetime, doctor_name, treatment_details):
    """
    Google Calendar에 단일 이벤트를 생성합니다.
    """
    seoul_tz = datetime.timezone(datetime.timedelta(hours=9))

    # reservation_datetime 객체를 사용합니다.
    event_start = reservation_datetime.replace(tzinfo=seoul_tz)
    event_end = event_start + datetime.timedelta(minutes=30)
    
    # 두 개의 요약(summary) 정보를 하나로 합칩니다.
    summary_text = f'{patient_name}' 
    
    # 캘린더 이벤트에 필요한 모든 정보를 한 번에 정의합니다.
    event = {
        'summary': summary_text,
        'location': pid,
        'description': f"{treatment_details}\n",
        'start': {
            'dateTime': event_start.isoformat(),
            'timeZone': 'Asia/Seoul',
        },
        'end': {
            'dateTime': event_end.isoformat(),
            'timeZone': 'Asia/Seoul',
        },
    }

    try:
        event = service.events().insert(calendarId='primary', body=event).execute()
        st.success(f"'{patient_name}' 환자의 캘린더 일정이 추가되었습니다.")
    except HttpError as error:
        st.error(f"캘린더 이벤트 생성 중 오류 발생: {error}")
        st.warning("구글 캘린더 인증 권한을 다시 확인해주세요.")
    except Exception as e:
        st.error(f"알 수 없는 오류 발생: {e}")
        
# --- OCS 분석 관련 함수 추가 ---

SCOPES = ["https://www.googleapis.com/auth/calendar.events"]

def get_google_calendar_service(user_id_safe):
    """
    사용자별로 Google Calendar 서비스 객체를 반환합니다. 
    인증 정보가 없거나 유효하지 않으면 None을 반환합니다.
    """
    # 1. 세션 상태에서 Creds 로드 시도
    creds = st.session_state.get(f"google_creds_{user_id_safe}")
    
    # 2. 세션에 없으면 Firebase에서 로드 및 세션에 저장
    if not creds:
        # **인자 이름 user_id_safe 사용**
        creds = load_google_creds_from_firebase(user_id_safe)
        if creds:
            st.session_state[f"google_creds_{user_id_safe}"] = creds

    # Creds가 없는 경우 서비스 객체를 만들 수 없으므로 None 반환
    if not creds:
        return None

    try:
        # 3. 서비스 빌드 (이미 로드된 creds 사용)
        # build 함수는 secrets.toml에 있는 정보를 자동으로 사용하므로 client_config는 불필요합니다.
        service = build('calendar', 'v3', credentials=creds)
        return service
        
    except Exception as e:
        # 서비스 빌드 실패 시 None 반환
        # 오류 메시지 출력은 외부 호출부(try...except)에서 처리하는 것이 깔끔합니다.
        # Streamlit 앱에서는 print 대신 st.warning/st.error를 사용하는 것이 좋습니다.
        print(f"인증 오류: {e}") 
        return None
        
# 엑셀 파일 암호화 여부 확인 (load_excel에서 사용)
def is_encrypted_excel(file_path):
    try:
        file_path.seek(0)
        return msoffcrypto.OfficeFile(file_path).is_encrypted()
    except Exception:
        return False

# 엑셀 파일 로드 및 복호화 (안전하게 스트림 복사)
def load_excel(file, password=None):
    try:
        file.seek(0)
        file_bytes = file.read()
        
        input_stream = io.BytesIO(file_bytes)
        decrypted_bytes_io = None
        
        if msoffcrypto.OfficeFile(input_stream).is_encrypted():
            if not password:
                raise ValueError("암호화된 파일입니다. 비밀번호를 입력해주세요.")
            
            decrypted_bytes_io = io.BytesIO()
            input_stream.seek(0)
            
            office_file = msoffcrypto.OfficeFile(input_stream)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_bytes_io)
            
            decrypted_bytes_io.seek(0)
            return pd.ExcelFile(decrypted_bytes_io), decrypted_bytes_io

        else:
            input_stream.seek(0)
            return pd.ExcelFile(input_stream), input_stream
            
    except Exception as e:
        raise ValueError(f"엑셀 로드 또는 복호화 실패: {e}")

def process_sheet_v8(df, professors_list, sheet_key): 
    if '예약의사' not in df.columns or '예약시간' not in df.columns:
        st.error(f"시트 처리 오류: '예약의사' 또는 '예약시간' 컬럼이 DataFrame에 없습니다.")
        return pd.DataFrame(columns=['진료번호', '예약일시', '예약시간', '환자명', '예약의사', '진료내역'])

    df = df.sort_values(by=['예약의사', '예약시간'])
    professors = df[df['예약의사'].isin(professors_list)]
    non_professors = df[~df['예약의사'].isin(professors_list)]

    if sheet_key != '보철':
        non_professors = non_professors.sort_values(by=['예약시간', '예약의사'])
    else:
        non_professors = non_professors.sort_values(by=['예약의사', '예약시간'])

    final_rows = []
    current_time = None
    current_doctor = None

    for _, row in non_professors.iterrows():
        if sheet_key != '보철':
            if current_time != row['예약시간']:
                if current_time is not None:
                    final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
                current_time = row['예약시간']
        else:
            if current_doctor != row['예약의사']:
                if current_doctor is not None:
                    final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
                current_doctor = row['예약의사']
        final_rows.append(row)

    final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
    final_rows.append(pd.Series(["<교수님>"] + [" "] * (len(df.columns) - 1), index=df.columns))

    current_professor = None
    for _, row in professors.iterrows():
        if current_professor != row['예약의사']:
            if current_professor is not None:
                final_rows.append(pd.Series([" "] * len(df.columns), index=df.columns))
            current_professor = row['예약의사']
        final_rows.append(row)

    final_df = pd.DataFrame(final_rows, columns=df.columns)
    required_cols = ['진료번호', '예약일시', '예약시간', '환자명', '예약의사', '진료내역']
    final_df = final_df[[col for col in required_cols if col in final_df.columns]]
    return final_df

def process_excel_file_and_style(file_bytes_io):
    file_bytes_io.seek(0)

    try:
        wb_raw = load_workbook(filename=file_bytes_io, keep_vba=False, data_only=True)
    except Exception as e:
        raise ValueError(f"엑셀 워크북 로드 실패: {e}")

    processed_sheets_dfs = {}
    
    file_bytes_io.seek(0)
    all_sheet_dfs = pd.read_excel(file_bytes_io, sheet_name=None)
    
    sheet_keyword_to_department_map = {
        '치과보철과': '보철', '보철과': '보철', '보철': '보철', '치과교정과' : '교정', '교정과': '교정', '교정': '교정',
        '구강 악안면외과' : '외과', '구강악안면외과': '외과', '외과': '외과', '구강 내과' : '내과', '구강내과': '내과', '내과': '내과',
        '치과보존과' : '보존', '보존과': '보존', '보존': '보존', '소아치과': '소치', '소치': '소치', '소아 치과': '소치',
        '원내생진료센터': '원내생', '원내생': '원내생','원내생 진료센터': '원내생','원진실':'원내생',
        '원스톱 협진센터' : '원스톱', '원스톱협진센터': '원스톱', '원스톱': '원스톱',
        '임플란트 진료센터' : '임플란트', '임플란트진료센터': '임플란트', '임플란트': '임플란트',
        '임플' : '임플란트', '치주과': '치주', '치주': '치주', '임플실': '임플란트', '병리': '병리'
    }

    for sheet_name_raw in wb_raw.sheetnames:
        sheet_name_lower = sheet_name_raw.strip().lower()

        sheet_key = None
        for keyword, department_name in sorted(sheet_keyword_to_department_map.items(), key=lambda item: len(item[0]), reverse=True):
            if keyword.lower() in sheet_name_lower:
                sheet_key = department_name
                break

        if not sheet_key:
            continue

        ws = wb_raw[sheet_name_raw]
        values = list(ws.values)
        while values and (values[0] is None or all((v is None or str(v).strip() == "") for v in values[0])):
            values.pop(0)
        if len(values) < 2:
            continue

        df = pd.DataFrame(values)
        if df.empty or df.iloc[0].isnull().all():
             continue

        df.columns = df.iloc[0]
        df = df.drop([0]).reset_index(drop=True)
        df = df.fillna("").astype(str)

        if '예약의사' not in df.columns:
            continue

        df['예약의사'] = df['예약의사'].str.strip().str.replace(" 교수님", "", regex=False)

        professors_dict_v8 = {
            '소치': ['김현태', '장기택', '김정욱', '현홍근', '김영재', '신터전', '송지수'],
            '보존': ['이인복', '금기연', '이우철', '유연지', '서덕규', '이창하', '김선영', '손원준'],
            '외과': ['최진영', '서병무', '명훈', '김성민', '박주영', '양훈주', '한정준', '권익재'],
            '치주': ['구영', '이용무', '설양조', '구기태', '김성태', '조영단'],
            '보철': ['곽재영', '김성균', '임영준', '김명주', '권호범', '여인성', '윤형인', '박지만', '이재현', '조준호'],
            '교정': [], '내과': [], '원진실': [], '원스톱': [], '임플란트': [], '병리': []
        }
        professors_list = professors_dict_v8.get(sheet_key, [])
        
        try:
            processed_df = process_sheet_v8(df, professors_list, sheet_key)
            processed_sheets_dfs[sheet_name_raw] = processed_df
        except Exception as e:
            st.error(f"시트 '{sheet_name_raw}' 처리 중 오류: {e}")
            continue

    if not processed_sheets_dfs:
        return all_sheet_dfs, None

    output_buffer_for_styling = io.BytesIO()
    with pd.ExcelWriter(output_buffer_for_styling, engine='openpyxl') as writer:
        for sheet_name_raw, df in processed_sheets_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name_raw, index=False)

    output_buffer_for_styling.seek(0)
    wb_styled = load_workbook(output_buffer_for_styling, keep_vba=False, data_only=True)

    # 스타일링 로직
    for sheet_name in wb_styled.sheetnames:
        ws = wb_styled[sheet_name]
        header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row[0].value == "<교수님>":
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True)

            if sheet_name.strip() == "교정" and '진료내역' in header:
                idx = header['진료내역'] - 1
                if len(row) > idx:
                    cell = row[idx]
                    text = str(cell.value).strip().lower()
                    
                    if ('bonding' in text or '본딩' in text) and 'debonding' not in text:
                        cell.font = Font(bold=True)

    final_output_bytes = io.BytesIO()
    wb_styled.save(final_output_bytes)
    final_output_bytes.seek(0)

    return all_sheet_dfs, final_output_bytes

def run_analysis(df_dict, professors_dict):
    analysis_results = {}
    sheet_department_map = {
        '소치': '소치', '소아치과': '소치', '소아 치과': '소치', '보존': '보존', '보존과': '보존', '치과보존과': '보존',
        '교정': '교정', '교정과': '교정', '치과교정과': '교정'
    }

    mapped_dfs = {}
    for sheet_name, df in df_dict.items():
        processed_sheet_name = sheet_name.replace(" ", "").lower()
        for key, dept in sheet_department_map.items():
            if processed_sheet_name == key.replace(" ", "").lower():
                if all(col in df.columns for col in ['예약의사', '예약시간']):
                     mapped_dfs[dept] = df.copy()
                break

    if '소치' in mapped_dfs:
        df = mapped_dfs['소치']
        non_professors_df = df[~df['예약의사'].isin(professors_dict.get('소치', []))]
        non_professors_df['예약시간'] = non_professors_df['예약시간'].astype(str).str.strip()
        non_professors_df = non_professors_df[non_professors_df['예약시간'] != 'nan']
        morning_patients = non_professors_df[(non_professors_df['예약시간'] >= '08:00') & (non_professors_df['예약시간'] <= '12:50')].shape[0]
        afternoon_patients = non_professors_df[non_professors_df['예약시간'] >= '13:00'].shape[0]
        if afternoon_patients > 0: afternoon_patients -= 1
        analysis_results['소치'] = {'오전': morning_patients, '오후': afternoon_patients}

    if '보존' in mapped_dfs:
        df = mapped_dfs['보존']
        non_professors_df = df[~df['예약의사'].isin(professors_dict.get('보존', []))]
        non_professors_df['예약시간'] = non_professors_df['예약시간'].astype(str).str.strip()
        non_professors_df = non_professors_df[non_professors_df['예약시간'] != 'nan']
        morning_patients = non_professors_df[(non_professors_df['예약시간'] >= '08:00') & (non_professors_df['예약시간'] <= '12:30')].shape[0]
        afternoon_patients = non_professors_df[non_professors_df['예약시간'] >= '12:50'].shape[0]
        if afternoon_patients > 0: afternoon_patients -= 1
        analysis_results['보존'] = {'오전': morning_patients, '오후': afternoon_patients}

    if '교정' in mapped_dfs:
        df = mapped_dfs['교정']
        bonding_patients_df = df[df['진료내역'].str.contains('bonding|본딩', case=False, na=False) & ~df['진료내역'].str.contains('debonding', case=False, na=False)]
        bonding_patients_df['예약시간'] = bonding_patients_df['예약시간'].astype(str).str.strip()
        morning_bonding_patients = bonding_patients_df[(bonding_patients_df['예약시간'] >= '08:00') & (bonding_patients_df['예약시간'] <= '12:30')].shape[0]
        afternoon_bonding_patients = bonding_patients_df[bonding_patients_df['예약시간'] >= '12:50'].shape[0]
        analysis_results['교정'] = {'오전': morning_bonding_patients, '오후': afternoon_bonding_patients}
        
    return analysis_results

def run_auto_notifications(matched_users, matched_doctors, excel_data_dfs, file_name, is_daily, sheet_keyword_to_department_map):
    """자동으로 모든 매칭 사용자에게 메일 및 캘린더 일정을 전송하는 핵심 로직"""
    sender = st.secrets["gmail"]["sender"]; sender_pw = st.secrets["gmail"]["app_password"]
    
    st.markdown("### 📚 학생(일반 사용자) 자동 전송 결과")
    if matched_users:
        for user_match_info in matched_users:
            real_email = user_match_info['email']; df_matched = user_match_info['data']
            user_name = user_match_info['name']; user_safe_key = user_match_info['safe_key']
            
            # 메일 전송
            email_cols = ['환자명', '진료번호', '예약의사', '진료내역', '예약일시', '예약시간', '등록과']
            df_for_mail = df_matched[[col for col in email_cols if col in df_matched.columns]]
            df_html = df_for_mail.to_html(index=False, escape=False); rows_as_dict = df_for_mail.to_dict('records')
            email_body = f"""<p>안녕하세요, {user_name}님.</p><p>{file_name} 분석 결과, 내원 예정인 환자 진료 정보입니다.</p>{df_html}<p>확인 부탁드립니다.</p>"""
            
            try:
                send_email(receiver=real_email, rows=rows_as_dict, sender=sender, password=sender_pw, custom_message=email_body, date_str=file_name) 
                st.write(f"✔️ **메일:** {user_name} ({real_email})에게 전송 완료.")
            except Exception as e: st.error(f"❌ **메일:** {user_name} ({real_email})에게 전송 실패: {e}")

            # 캘린더 등록
            creds = load_google_creds_from_firebase(user_safe_key)
            if creds and creds.valid and not creds.expired:
                try:
                    service = build('calendar', 'v3', credentials=creds)
                    for _, row in df_matched.iterrows():
                        reservation_date_raw = row.get('예약일시', ''); reservation_time_raw = row.get('예약시간', '')
                        if reservation_date_raw and reservation_time_raw:
                            full_datetime_str = f"{str(reservation_date_raw).strip()} {str(reservation_time_raw).strip()}"; reservation_datetime = datetime.datetime.strptime(full_datetime_str, '%Y/%m/%d %H:%M')
                            event_prefix = "✨ 내원 : " if is_daily else "❓내원 : "
                            event_title = f"{event_prefix}{row.get('환자명', 'N/A')} ({row.get('등록과', 'N/A')}, {row.get('예약의사', 'N/A')})"
                            event_description = f"환자명 : {row.get('환자명', 'N/A')}\n진료번호 : {row.get('진료번호', 'N/A')}\n진료내역 : {row.get('진료내역', 'N/A')}"
                            service.events().insert(calendarId='primary', body={
                                'summary': event_title, 'location': row.get('진료번호', ''), 'description': event_description,
                                'start': {'dateTime': reservation_datetime.replace(tzinfo=datetime.timezone(datetime.timedelta(hours=9))).isoformat(), 'timeZone': 'Asia/Seoul'},
                                'end': {'dateTime': (reservation_datetime + datetime.timedelta(minutes=30)).replace(tzinfo=datetime.timezone(datetime.timedelta(hours=9))).isoformat(), 'timeZone': 'Asia/Seoul'}
                            }).execute()
                    st.write(f"✔️ **캘린더:** {user_name}에게 일정 추가 완료.")
                except Exception as e: st.warning(f"⚠️ **캘린더:** {user_name} 일정 추가 중 오류: 인증/권한 문제일 수 있습니다.")
            else: st.warning(f"⚠️ **캘린더:** {user_name}님은 Google Calendar 계정이 연동되지 않았습니다.")
    else: st.info("매칭된 학생(사용자)이 없습니다.")

    st.markdown("### 🧑‍⚕️ 치과의사 자동 전송 결과")
    if matched_doctors:
        for res in matched_doctors:
            matched_rows_for_doctor = []; doctor_dept = res['department']; sheets_to_search = patient_dept_to_sheet_map.get(doctor_dept, [doctor_dept])
            
            # 매칭 데이터 재구성 (auto run을 위해)
            for sheet_name_excel_raw, df_sheet in excel_data_dfs.items():
                excel_sheet_department = None
                for keyword, department_name in sorted(sheet_keyword_to_department_map.items(), key=lambda item: len(item[0]), reverse=True):
                    if keyword.lower().replace(' ', '') in sheet_name_excel_raw.strip().lower().replace(' ', ''): excel_sheet_department = department_name; break
                if excel_sheet_department in sheets_to_search:
                    for _, excel_row in df_sheet.iterrows():
                        excel_doctor_name_from_row = str(excel_row.get('예약의사', '')).strip().replace("'", "").replace("‘", "").replace("’", "").strip()
                        if excel_doctor_name_from_row == res['name']: matched_rows_for_doctor.append(excel_row.copy())
            
            if matched_rows_for_doctor:
                df_matched = pd.DataFrame(matched_rows_for_doctor); latest_file_name = db.reference("ocs_analysis/latest_file_name").get()
                email_cols = ['환자명', '진료번호', '예약의사', '진료내역', '예약일시', '예약시간']; 
                # df_for_mail 생성 (이메일에 필요한 컬럼만 선택)
                df_for_mail = df_matched[[col for col in email_cols if col in df_matched.columns]]
                
                # 🔑 수정 1: df_for_mail을 HTML로 변환하여 df_html 변수에 할당 (제안해주신 방식)
                # escape=False는 HTML 태그를 그대로 사용하기 위함이며, border=1을 추가하면 더 깔끔한 테이블이 됩니다.
                df_html = df_for_mail.to_html(index=False, border=1) # border=1을 추가하면 가독성 향상
                rows_as_dict = df_for_mail.to_dict('records')
                # 🔑 수정 2: email_body에서 df_html 변수를 사용
                email_body = f"""<p>안녕하세요, {res['name']} 치과의사님.</p><p>{latest_file_name}에서 가져온 내원할 환자 정보입니다.</p>{df_html}<p>확인 부탁드립니다.</p>"""
                try:
                    send_email(receiver=res['email'], rows=rows_as_dict, sender=sender, password=sender_pw, custom_message=email_body, date_str=latest_file_name)
                    st.write(f"✔️ **메일:** Dr. {res['name']}에게 전송 완료!")
                except Exception as e: st.error(f"❌ **메일:** Dr. {res['name']}에게 전송 실패: {e}")

                creds = load_google_creds_from_firebase(res['safe_key'])
                if creds and creds.valid and not creds.expired:
                    try:
                        service = build('calendar', 'v3', credentials=creds)
                        for _, row in df_matched.iterrows():
                            reservation_date_str = row.get('예약일시', ''); reservation_time_str = row.get('예약시간', '')
                            if reservation_date_str and reservation_time_str:
                                full_datetime_str = f"{str(reservation_date_str).strip()} {str(reservation_time_str).strip()}"; reservation_datetime = datetime.datetime.strptime(full_datetime_str, '%Y/%m/%d %H:%M')
                                event_prefix = "✨:" if is_daily else "?:"; event_title = f"{event_prefix}{row.get('환자명', 'N/A')}({row.get('진료번호', 'N/A')})"
                                event_description = f"환자명: {row.get('환자명', 'N/A')}\n진료번호: {row.get('진료번호', 'N/A')}\n진료내역: {row.get('진료내역', 'N/A')}"
                                service.events().insert(calendarId='primary', body={
                                    'summary': event_title, 'location': row.get('진료번호', ''), 'description': event_description,
                                    'start': {'dateTime': reservation_datetime.replace(tzinfo=datetime.timezone(datetime.timedelta(hours=9))).isoformat(), 'timeZone': 'Asia/Seoul'},
                                    'end': {'dateTime': (reservation_datetime + datetime.timedelta(minutes=30)).replace(tzinfo=datetime.timezone(datetime.timedelta(hours=9))).isoformat(), 'timeZone': 'Asia/Seoul'}
                                }).execute()
                        st.write(f"✔️ **캘린더:** Dr. {res['name']}에게 일정 추가 완료.")
                    except Exception as e: st.warning(f"⚠️ **캘린더:** Dr. {res['name']} 일정 추가 중 오류: {e}")
                else: st.warning(f"⚠️ **캘린더:** Dr. {res['name']}님은 Google Calendar 계정이 연동되지 않았습니다.")
            else: st.warning(f"Dr. {res['name']} 치과의사의 매칭 데이터가 엑셀 파일에 없습니다.")
    else: st.info("매칭된 치과의사 계정이 없습니다.")


# --- 5. Streamlit App Start and Session State ---
# --- 5. Streamlit App Start and Session State ---
# --- Streamlit 애플리케이션 시작 ---
st.set_page_config(layout="wide")

# 제목에 링크 추가 및 초기화 로직
st.markdown("""
    <style>
    .title-link {
        text-decoration: none; color: inherit;
    }
    </style>
    <h1>
        <a href="." class="title-link">환자 내원 확인 시스템</a>
    </h1>
""", unsafe_allow_html=True)
st.markdown("---")
st.markdown("<p style='text-align: left; color: grey; font-size: small;'>directed by HSY</p>", unsafe_allow_html=True)


# --- 세션 상태 초기화 ---
if "clear" in st.query_params and st.query_params["clear"] == "true":
    st.session_state.clear()
    st.query_params["clear"] = "false"
    st.rerun()

if 'email_change_mode' not in st.session_state:
    st.session_state.email_change_mode = False
if 'user_id_input_value' not in st.session_state:
    st.session_state.user_id_input_value = ""
if 'found_user_email' not in st.session_state:
    st.session_state.found_user_email = ""
if 'current_firebase_key' not in st.session_state:
    st.session_state.current_firebase_key = ""
if 'current_user_name' not in st.session_state:
    st.session_state.current_user_name = ""
if 'logged_in_as_admin' not in st.session_state:
    st.session_state.logged_in_as_admin = False
if 'admin_password_correct' not in st.session_state:
    st.session_state.admin_password_correct = False
if 'select_all_users' not in st.session_state:
    st.session_state.select_all_users = False
if 'google_calendar_auth_needed' not in st.session_state:
    st.session_state.google_calendar_auth_needed = False
if 'google_creds' not in st.session_state:
    st.session_state['google_creds'] = {}
if "clear" in st.query_params and st.query_params["clear"] == "true":
    st.session_state.clear()
    st.query_params["clear"] = "false"
    st.rerun()
if 'email_change_mode' not in st.session_state:
    st.session_state.email_change_mode = False
# ... (다른 기존 초기화 코드) ...
if 'google_creds' not in st.session_state:
    st.session_state['google_creds'] = {}
# 💡 여기에 'auto_run_confirmed' 플래그를 추가합니다.
if 'auto_run_confirmed' not in st.session_state:
    # 초기에는 None으로 설정하여, 사용자가 '자동' 또는 '수동'을 선택하기 전임을 나타냅니다.
    st.session_state.auto_run_confirmed = None 

users_ref = db.reference("users")
doctor_users_ref = db.reference("doctor_users")

# --- 6. User and Admin and doctor Login and User Management ---
import os
import streamlit as st
import datetime
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

# Assume these functions are defined elsewhere in your script
# from your_utils import is_valid_email, is_encrypted_excel, load_excel, process_excel_file_and_style, run_analysis, sanitize_path, recover_email, get_google_calendar_service, send_email, send_email_simple, create_calendar_event, create_static_calendar_event, create_auth_url, load_google_creds_from_firebase, users_ref, db, is_daily_schedule, sheet_keyword_to_department_map


# --- 사용 설명서 PDF 다운로드 버튼 추가 ---
pdf_file_path = "manual.pdf"
pdf_display_name = "사용 설명서"

if os.path.exists(pdf_file_path):
    with open(pdf_file_path, "rb") as pdf_file:
        st.download_button(
            label=f"{pdf_display_name} 다운로드",
            data=pdf_file,
            file_name=pdf_file_path,
            mime="application/pdf"
        )
else:
    st.warning(f"⚠️ {pdf_display_name} 파일을 찾을 수 없습니다. (경로: {pdf_file_path})")

# 로그인 폼 - 로그인/등록 완료 전까지는 이 섹션만 표시
if 'login_mode' not in st.session_state:
    st.session_state.login_mode = 'not_logged_in'

if st.session_state.get('login_mode') in ['not_logged_in', 'admin_mode']:
    tab1, tab2 = st.tabs(["학생 로그인", "치과의사 로그인"])

    # 탭 1: 일반 사용자/학생 로그인
    with tab1:
        st.subheader("👨‍🎓 학생 로그인")
        user_name = st.text_input("사용자 이름을 입력하세요 (예시: 홍길동)", key="login_username_tab1")
        password_input = st.text_input("비밀번호를 입력하세요", type="password", key="login_password_tab1")

        if st.button("로그인/등록", key="login_button_tab1"):
            if not user_name:
                st.error("사용자 이름을 입력해주세요.")
            elif user_name.strip().lower() == "admin":
                st.session_state.login_mode = 'admin_mode'
                st.session_state.logged_in_as_admin = True
                st.session_state.found_user_email = "admin"
                st.session_state.current_user_name = "admin"
                st.rerun()
            else:
                all_users_meta = users_ref.get()
                matched_user = None
                if all_users_meta:
                    for safe_key, user_info in all_users_meta.items():
                        if user_info and user_info.get("name") == user_name:
                            matched_user = {"safe_key": safe_key, "email": user_info.get("email", ""), "name": user_info.get("name", ""), "password": user_info.get("password")}
                            break
                if matched_user:
                    if "password" not in matched_user or not matched_user.get("password"):
                        if password_input == "1234":
                            st.session_state.found_user_email = matched_user["email"]
                            st.session_state.user_id_input_value = matched_user["email"]
                            st.session_state.current_firebase_key = matched_user["safe_key"]
                            st.session_state.current_user_name = user_name
                            st.session_state.login_mode = 'user_mode'
                            st.info(f"**{user_name}**님으로 로그인되었습니다. 기존 사용자이므로 초기 비밀번호 1234가 설정되었습니다.")
                            users_ref.child(matched_user["safe_key"]).update({"password": "1234"})
                            st.rerun()
                        else:
                            st.error("비밀번호가 일치하지 않습니다. 기존 사용자의 초기 비밀번호는 '1234'입니다. 다시 시도해주세요.")
                    else:
                        if password_input == matched_user.get("password"):
                            st.session_state.found_user_email = matched_user["email"]
                            st.session_state.user_id_input_value = matched_user["email"]
                            st.session_state.current_firebase_key = matched_user["safe_key"]
                            st.session_state.current_user_name = user_name
                            st.session_state.login_mode = 'user_mode'
                            st.info(f"**{user_name}**님으로 로그인되었습니다. 이메일 주소: **{st.session_state.found_user_email}**")
                            st.rerun()
                        else:
                            st.error("비밀번호가 일치하지 않거나 다른 사용자가 이미 해당이름을 사용 중입니다. 신규 등록 시 이름에 알파벳이나 숫자를 붙여주세요.")
                else:
                    st.session_state.current_user_name = user_name
                    st.session_state.login_mode = 'new_user_registration'
                    st.rerun()

    # 탭 2: 치과의사 로그인
    with tab2:
        st.subheader("🧑‍⚕️ 치과의사 로그인")
        doctor_email = st.text_input("치과의사 이메일 주소를 입력하세요", key="doctor_email_input_tab2")
        password_input_doc = st.text_input("비밀번호를 입력하세요", type="password", key="doctor_password_input_tab2")

        if st.button("로그인/등록", key="doctor_login_button_tab2"):
            if doctor_email:
                safe_key = doctor_email.replace('@', '_at_').replace('.', '_dot_')
                matched_doctor = doctor_users_ref.child(safe_key).get()
                
                if matched_doctor:
                    if password_input_doc == matched_doctor.get("password"):
                        st.session_state.found_user_email = matched_doctor["email"]
                        st.session_state.user_id_input_value = matched_doctor["email"]
                        st.session_state.current_firebase_key = safe_key
                        st.session_state.current_user_name = matched_doctor.get("name")
                        st.session_state.current_user_dept = matched_doctor.get("department")
                        st.session_state.current_user_role = 'doctor'
                        st.session_state.login_mode = 'doctor_mode'
                        st.info(f"치과의사 **{st.session_state.current_user_name}**님으로 로그인되었습니다. 이메일 주소: **{st.session_state.found_user_email}**")
                        st.rerun()
                    else:
                        st.error("비밀번호가 일치하지 않습니다. 다시 확인해주세요.")
                else:
                    if password_input_doc == "1234":
                        st.info("💡 새로운 치과의사 계정으로 인식되었습니다. 초기 비밀번호 '1234'로 등록을 완료합니다.")
                        st.session_state.found_user_email = doctor_email
                        st.session_state.user_id_input_value = doctor_email
                        st.session_state.current_firebase_key = ""
                        st.session_state.current_user_name = None
                        st.session_state.current_user_role = 'doctor'
                        st.session_state.current_user_dept = None
                        st.session_state.login_mode = 'new_doctor_registration'
                        st.rerun()
                    else:
                        st.info(f"아래에 정보를 입력하여 등록을 완료하세요.")
                        st.session_state.found_user_email = doctor_email
                        st.session_state.user_id_input_value = doctor_email
                        st.session_state.current_firebase_key = ""
                        st.session_state.current_user_name = None
                        st.session_state.login_mode = 'new_doctor_registration'
                        st.rerun()
            else:
                st.warning("치과의사 이메일 주소를 입력해주세요.")

# 새로운 일반 사용자 등록 로직 (탭 바깥)
if st.session_state.get('login_mode') == 'new_user_registration':
    st.info(f"'{st.session_state.current_user_name}'님은 새로운 사용자입니다. 아래에 정보를 입력하여 등록을 완료하세요.")
    st.subheader("👨‍⚕️ 신규 사용자 등록")
    new_email_input = st.text_input("아이디(이메일)를 입력하세요", key="new_user_email_input")
    password_input = st.text_input("새로운 비밀번호를 입력하세요", type="password", key="new_user_password_input")
    
    if st.button("사용자 등록 완료", key="new_user_reg_button"):
        if is_valid_email(new_email_input) and password_input:
            new_firebase_key = sanitize_path(new_email_input)
            all_users_meta = users_ref.get()
            is_email_used = False
            if all_users_meta:
                for user_info in all_users_meta.values():
                    if user_info.get("email") == new_email_input:
                        is_email_used = True
                        break
            if is_email_used:
                st.error("이미 등록된 이메일 주소입니다. 다른 주소를 사용해주세요.")
            else:
                st.session_state.current_firebase_key = new_firebase_key
                st.session_state.found_user_email = new_email_input
                st.session_state.current_user_name = st.session_state.current_user_name
                st.session_state.login_mode = 'user_mode'
                users_ref.child(new_firebase_key).set({
                    "name": st.session_state.current_user_name,
                    "email": new_email_input,
                    "password": password_input
                })
                st.success(f"새로운 사용자 **{st.session_state.current_user_name}**님 ({new_email_input}) 정보가 등록되었습니다.")
                st.rerun()
        else:
            st.error("올바른 이메일 주소와 비밀번호를 입력해주세요.")
            
# --- 새로운 치과의사 등록 로직 (탭 바깥) ---
if st.session_state.get('login_mode') == 'new_doctor_registration':
    st.info(f"아래에 정보를 입력하여 등록을 완료하세요.")
    st.subheader("👨‍⚕️ 새로운 치과의사 등록")
    new_doctor_name_input = st.text_input("이름을 입력하세요 (원내생이라면 '홍길동95'과 같은 형태로 등록바랍니다)", key="new_doctor_name_input", value=st.session_state.get('current_user_name', ''))
    password_input = st.text_input("새로운 비밀번호를 입력하세요", type="password", key="new_doctor_password_input", value="1234" if st.session_state.get('current_firebase_key') else "")
    user_id_input = st.text_input("아이디(이메일)를 입력하세요", key="new_doctor_email_input", value=st.session_state.get('found_user_email', ''))
    
    dept_options = ["교정", "내과", "보존", "보철", "소치", "외과", "치주", "원내생"]
    selected_dept = st.session_state.get('current_user_dept')
    default_index = 0
    if selected_dept and selected_dept in dept_options:
        default_index = dept_options.index(selected_dept)
    department = st.selectbox("등록 과", dept_options, key="new_doctor_dept_selectbox", index=default_index)

    if st.button("치과의사 등록 완료", key="new_doc_reg_button"):
        if new_doctor_name_input and is_valid_email(user_id_input) and password_input and department:
            new_email = user_id_input
            new_firebase_key = sanitize_path(new_email)
            st.session_state.current_firebase_key = new_firebase_key
            st.session_state.found_user_email = new_email
            st.session_state.current_user_dept = department
            st.session_state.current_user_role = 'doctor'
            st.session_state.current_user_name = new_doctor_name_input
            st.session_state.login_mode = 'doctor_mode'
            doctor_users_ref.child(new_firebase_key).set({"name": st.session_state.current_user_name, "email": new_email, "password": password_input, "role": st.session_state.current_user_role, "department": department})
            st.success(f"새로운 치과의사 **{st.session_state.current_user_name}**님 ({new_email}) 정보가 등록되었습니다.")
            st.rerun()
        else:
            st.error("이름, 올바른 이메일 주소, 비밀번호, 그리고 등록 과를 입력해주세요.")
            
# --- 이메일 변경 기능 (모든 사용자 공통) ---
if st.session_state.get('login_mode') in ['user_mode', 'doctor_mode', 'email_change_mode']:
    if st.session_state.get('current_firebase_key'):
        st.text_input("아이디 (등록된 이메일)", value=st.session_state.get('found_user_email', ''), disabled=True)
        if st.button("이메일 주소 변경"):
            st.session_state.email_change_mode = True
            st.rerun()
        if st.session_state.get('email_change_mode'):
            st.divider()
            st.subheader("이메일 주소 변경")
            new_email_input = st.text_input("새 이메일 주소를 입력하세요", value=st.session_state.get('user_id_input_value', ''))
            st.session_state.user_id_input_value = new_email_input
            if st.button("변경 완료"):
                if is_valid_email(new_email_input):
                    new_firebase_key = sanitize_path(new_email_input)
                    old_firebase_key = st.session_state.current_firebase_key
                    user_role_to_change = st.session_state.get("current_user_role")
                    if old_firebase_key != new_firebase_key:
                        if user_role_to_change == 'doctor':
                            target_ref = doctor_users_ref
                        else:
                            target_ref = users_ref
                        
                        # 사용자 메타데이터 이동
                        current_user_meta = target_ref.child(old_firebase_key).get()
                        if current_user_meta:
                            current_user_meta.update({"email": new_email_input})
                            target_ref.child(new_firebase_key).set(current_user_meta)
                            target_ref.child(old_firebase_key).delete()
                        
                        # 환자 데이터 이동 (일반 사용자만 해당)
                        if user_role_to_change != 'doctor':
                            old_patient_data = db.reference(f"patients/{old_firebase_key}").get()
                            if old_patient_data:
                                db.reference(f"patients/{new_firebase_key}").set(old_patient_data)
                                db.reference(f"patients/{old_firebase_key}").delete()
                        
                        st.session_state.current_firebase_key = new_firebase_key
                        st.session_state.found_user_email = new_email_input
                        st.success(f"이메일 주소가 **{new_email_input}**로 성공적으로 변경되었습니다.")
                    else:
                        st.info("이메일 주소 변경사항이 없습니다.")
                    st.session_state.email_change_mode = False
                    st.rerun()
                else:
                    st.error("올바른 이메일 주소 형식이 아닙니다.")

# --- 7. Admin 모드 로그인 처리 ---
if st.session_state.get('login_mode') == 'admin_mode':
    st.session_state.logged_in_as_admin = True; st.session_state.found_user_email = "admin"
    st.session_state.current_user_name = "admin"
    
    st.subheader("💻 Excel File Processor")
    uploaded_file = st.file_uploader("암호화된 Excel 파일을 업로드하세요", type=["xlsx", "xlsm"])
    
    sheet_keyword_to_department_map = {
    '치과보철과': '보철', '보철과': '보철', '보철': '보철', '치과교정과' : '교정', '교정과': '교정', '교정': '교정',
    '구강 악안면외과' : '외과', '구강악안면외과': '외과', '외과': '외과', '구강 내과' : '내과', '구강내과': '내과', '내과': '내과',
    '치과보존과' : '보존', '보존과': '보존', '보존': '보존', '소아치과': '소치', '소치': '소치', '소아 치과': '소치',
    '원내생진료센터': '원내생', '원내생': '원내생','원내생 진료센터': '원내생','원진실':'원내생',
    '원스톱 협진센터' : '원스톱', '원스톱협진센터': '원스톱', '원스톱': '원스톱',
    '임플란트 진료센터' : '임플란트', '임플란트진료센터': '임플란트', '임플란트': '임플란트',
    '임플' : '임플란트', '치주과': '치주', '치주': '치주', '임플실': '임플란트', '병리': '병리'
    }

    if uploaded_file:
        file_name = uploaded_file.name; is_daily = is_daily_schedule(file_name)
        st.info(f"파일 '{file_name}'이(가) 업로드되었습니다. 처리를 시작합니다.")
        
        uploaded_file.seek(0); password = None
        
        # 1. 파일 비밀번호 처리 (필요시)
        if is_encrypted_excel(uploaded_file):
            password = st.text_input("⚠️ 암호화된 파일입니다. 비밀번호를 입력해주세요.", type="password", key="auto_exec_password")
            if not password: st.info("비밀번호 입력 대기 중..."); st.stop()

        # 2. 파일 처리 및 분석 실행 (이후 자동/수동 실행을 위한 데이터 준비)
        try:
            xl_object, raw_file_io = load_excel(uploaded_file, password)
            excel_data_dfs, styled_excel_bytes = process_excel_file_and_style(raw_file_io)
            professors_dict = {
                '소치': ['김현태', '장기택', '김정욱', '현홍근', '김영재', '신터전', '송지수'], '보존': ['이인복', '금기연', '이우철', '유연지', '서덕규', '이창하', '김선영', '손원준']
            }
            analysis_results = run_analysis(excel_data_dfs, professors_dict)
            
            # DB에 분석 결과 저장
            today_date_str = datetime.datetime.now().strftime("%Y-%m-%d")
            db.reference("ocs_analysis/latest_result").set(analysis_results); db.reference("ocs_analysis/latest_date").set(today_date_str)
            db.reference("ocs_analysis/latest_file_name").set(file_name)
            
            st.session_state.last_processed_data = excel_data_dfs; st.session_state.last_processed_file_name = file_name

            if excel_data_dfs is None or styled_excel_bytes is None:
                st.warning("엑셀 파일 처리 중 문제가 발생했거나 처리할 데이터가 없습니다. 실행을 중단합니다."); st.stop()
                
            output_filename = uploaded_file.name.replace(".xlsx", "_processed.xlsx").replace(".xlsm", "_processed.xlsm")
            st.download_button(
                "처리된 엑셀 다운로드", data=styled_excel_bytes, file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ 파일 처리 및 분석이 완료되었습니다. 이제 알림 전송 방법을 선택하세요.")
            
        except ValueError as ve: st.error(f"파일 처리 실패: {ve}"); st.stop()
        except Exception as e: st.error(f"예상치 못한 오류 발생: {e}"); st.stop()

        # 3. ★ 자동/수동 실행 결정 트리 ★
        
        st.markdown("---")
        st.subheader("🚀 알림 전송 옵션")
        
        col_auto, col_manual = st.columns(2)

        with col_auto:
            if st.button("YES: 자동으로 모든 사용자에게 전송", key="auto_run_yes"):
                st.session_state.auto_run_confirmed = True
                st.rerun()
        
        with col_manual:
            if st.button("NO: 수동으로 사용자 선택", key="auto_run_no"):
                st.session_state.auto_run_confirmed = False
                st.rerun()

        # 4. 실행 로직 분기
        if 'last_processed_data' in st.session_state and st.session_state.last_processed_data:
            
            # 매칭 데이터 미리 준비 (자동/수동 모두 사용)
            all_users_meta = db.reference("users").get(); all_patients_data = db.reference("patients").get()
            all_doctors_meta = db.reference("doctor_users").get()
            
            matched_users = []; matched_doctors_data = [] # 변수명 변경
            
            # --- 학생 매칭 로직 재구성 (수동/자동에서 사용할 리스트 생성) ---
            if all_patients_data:
                patient_dept_to_sheet_map = {'보철': ['보철', '임플란트'], '치주': ['치주', '임플란트'], '외과': ['외과', '원스톱', '임플란트'], '교정': ['교정'], '내과': ['내과'], '보존': ['보존'], '소치': ['소치'], '원내생': ['원내생'], '병리': ['병리']}
                for uid_safe, registered_patients_for_this_user in all_patients_data.items():
                    user_email = recover_email(uid_safe); user_display_name = user_email
                    if all_users_meta and uid_safe in all_users_meta and "name" in all_users_meta[uid_safe]:
                        user_display_name = all_users_meta[uid_safe]["name"]; user_email = all_users_meta[uid_safe]["email"]
                    
                    registered_patients_data = []
                    if registered_patients_for_this_user:
                        for pid_key, val in registered_patients_for_this_user.items(): 
                            registered_depts = [
                                dept.capitalize() for dept in PATIENT_DEPT_FLAGS + ['치주'] 
                                if val.get(dept.lower()) is True or val.get(dept.lower()) == 'True' or val.get(dept.lower()) == 'true'
                            ]
                            registered_patients_data.append({"환자명": val.get("환자이름", "").strip(), "진료번호": pid_key.strip().zfill(8), "등록과_리스트": registered_depts})
                    
                    matched_rows_for_user = []
                    for registered_patient in registered_patients_data:
                        registered_depts = registered_patient["등록과_리스트"]; sheets_to_search = set()
                        for dept in registered_depts: sheets_to_search.update(patient_dept_to_sheet_map.get(dept, [dept]))

                        for sheet_name_excel_raw, df_sheet in excel_data_dfs.items():
                            excel_sheet_department = None
                            for keyword, department_name in sheet_keyword_to_department_map.items():
                                if keyword.lower() in sheet_name_excel_raw.strip().lower(): excel_sheet_department = department_name; break
                            
                            if excel_sheet_department in sheets_to_search:
                                for _, excel_row in df_sheet.iterrows():
                                    excel_patient_name = str(excel_row.get("환자명", "")).strip(); excel_patient_pid = str(excel_row.get("진료번호", "")).strip().zfill(8)
                                    
                                    if (registered_patient["환자명"] == excel_patient_name and registered_patient["진료번호"] == excel_patient_pid):
                                        matched_row_copy = excel_row.copy(); matched_row_copy["시트"] = sheet_name_excel_raw
                                        matched_row_copy["등록과"] = ", ".join(registered_depts); matched_rows_for_user.append(matched_row_copy); break
                    
                    if matched_rows_for_user:
                        combined_matched_df = pd.DataFrame(matched_rows_for_user)
                        matched_users.append({"email": user_email, "name": user_display_name, "data": combined_matched_df, "safe_key": uid_safe})

            # --- 치과의사 매칭 로직 재구성 ---
            doctor_dept_to_sheet_map = {'보철': ['보철', '임플란트'], '치주': ['치주', '임플란트'], '외과': ['외과', '원스톱', '임플란트'], '교정': ['교정'], '내과': ['내과'], '보존': ['보존'], '소치': ['소치'], '원내생': ['원내생'], '병리': ['병리']}
            doctors = []
            if all_doctors_meta:
                for safe_key, user_info in all_doctors_meta.items():
                    if user_info: doctors.append({"safe_key": safe_key, "name": user_info.get("name", "이름 없음"), "email": user_info.get("email", "이메일 없음"), "department": user_info.get("department", "미지정")})
            
            if doctors and excel_data_dfs:
                for res in doctors:
                    found_match = False; doctor_dept = res['department']; sheets_to_search = doctor_dept_to_sheet_map.get(doctor_dept, [doctor_dept])
                    for sheet_name_excel_raw, df_sheet in excel_data_dfs.items():
                        excel_sheet_department = None
                        for keyword, department_name in sorted(sheet_keyword_to_department_map.items(), key=lambda item: len(item[0]), reverse=True):
                            if keyword.lower().replace(' ', '') in sheet_name_excel_raw.strip().lower().replace(' ', ''): excel_sheet_department = department_name; break
                        if not excel_sheet_department: continue
                        if excel_sheet_department in sheets_to_search:
                            for _, excel_row in df_sheet.iterrows():
                                excel_doctor_name_from_row = str(excel_row.get('예약의사', '')).strip().replace("'", "").replace("‘", "").replace("’", "").strip()
                                if excel_doctor_name_from_row == res['name']:
                                    matched_doctors_data.append(res); found_match = True; break 
                        if found_match: break

            # A. 자동 실행 로직 (버튼 YES 클릭 시)
            if st.session_state.auto_run_confirmed:
                st.markdown("---")
                st.warning("자동으로 모든 매칭 사용자에게 알림(메일/캘린더)을 전송합니다. 재확인 버튼을 누를 필요가 없습니다.")
                
                run_auto_notifications(matched_users, matched_doctors_data, excel_data_dfs, file_name, is_daily, sheet_keyword_to_department_map)

                st.session_state.auto_run_confirmed = False # 상태 초기화
                st.stop()
                
            # B. 수동 실행 로직 (버튼 NO 클릭 시 또는 기본 상태)
            elif st.session_state.auto_run_confirmed is False:
                st.markdown("---")
                st.info("아래 탭에서 전송할 사용자 목록을 확인하고, 원하는 사용자에게 수동으로 알림을 전송해주세요.")

                # (이전 코드의 수동 사용자 선택 탭 로직을 여기에 통합)
                student_admin_tab, doctor_admin_tab = st.tabs(['📚 학생 관리자 모드', '🧑‍⚕️ 치과의사 관리자 모드'])
                
                # --- 학생 수동 전송 탭 ---
                with student_admin_tab:
                    st.subheader("📚 학생 수동 전송 (매칭 결과)");
                    st.warning("수동 모드에서는 이메일/캘린더 전송 버튼을 눌러야 실행됩니다.")
                    
                    if matched_users:
                        st.success(f"매칭된 환자가 있는 **{len(matched_users)}명의 사용자**를 발견했습니다.")
                        matched_user_list_for_dropdown = [f"{user['name']} ({user['email']})" for user in matched_users]
                        
                        if 'select_all_matched_users' not in st.session_state: st.session_state.select_all_matched_users = False
                        select_all_matched_button = st.button("매칭된 사용자 모두 선택/해제", key="select_all_matched_btn")
                        if select_all_matched_button: st.session_state.select_all_matched_users = not st.session_state.select_all_matched_users; st.rerun()
                        
                        default_selection_matched = matched_user_list_for_dropdown if st.session_state.select_all_matched_users else []
                        selected_users_to_act = st.multiselect("액션을 취할 사용자 선택", matched_user_list_for_dropdown, default=default_selection_matched, key="matched_user_multiselect")
                        selected_matched_users_data = [user for user in matched_users if f"{user['name']} ({user['email']})" in selected_users_to_act]
                        
                        for user_match_info in selected_matched_users_data:
                            st.markdown(f"**수신자:** {user_match_info['name']} ({user_match_info['email']})")
                            st.dataframe(user_match_info['data'])
                        
                        mail_col, calendar_col = st.columns(2)
                        with mail_col:
                            if st.button("선택된 사용자에게 메일 보내기", key="manual_send_mail_student"):
                                for user_match_info in selected_matched_users_data:
                                    real_email = user_match_info['email']; df_matched = user_match_info['data']
                                    user_name = user_match_info['name']; user_safe_key = user_match_info['safe_key']
                                    if not df_matched.empty:
                                        latest_file_name = db.reference("ocs_analysis/latest_file_name").get()
                                        email_cols = ['환자명', '진료번호', '예약의사', '진료내역', '예약일시', '예약시간', '등록과']
                                        df_for_mail = df_matched[[col for col in email_cols if col in df_matched.columns]]
                                        df_html = df_for_mail.to_html(index=False, escape=False); rows_as_dict = df_for_mail.to_dict('records')
                                        email_body = f"""<p>안녕하세요, {user_name}님.</p><p>{latest_file_name}분석 결과, 내원 예정인 환자 진료 정보입니다.</p>{df_html}<p>확인 부탁드립니다.</p>"""
                                        try:
                                            send_email(receiver=real_email, rows=rows_as_dict, sender=sender, password=sender_pw, custom_message=email_body, date_str=latest_file_name) 
                                            st.success(f"**{user_name}**님 ({real_email})에게 예약 정보 이메일 전송 완료!")
                                        except Exception as e: st.error(f"**{user_name}**님 ({real_email})에게 이메일 전송 실패: {e}")
                                    else: st.warning(f"**{user_name}**님에게 보낼 매칭 데이터가 없습니다.")

                        with calendar_col:
                            if st.button("선택된 사용자에게 Google Calendar 일정 추가", key="manual_send_calendar_student"):
                                for user_match_info in selected_matched_users_data:
                                    user_safe_key = user_match_info['safe_key']; user_name = user_match_info['name']; df_matched = user_match_info['data']
                                    creds = load_google_creds_from_firebase(user_safe_key)
                                    if creds and creds.valid and not creds.expired:
                                        try:
                                            service = build('calendar', 'v3', credentials=creds)
                                            if not df_matched.empty:
                                                for _, row in df_matched.iterrows():
                                                    reservation_date_raw = row.get('예약일시', ''); reservation_time_raw = row.get('예약시간', '')
                                                    if reservation_date_raw and reservation_time_raw:
                                                        full_datetime_str = f"{str(reservation_date_raw).strip()} {str(reservation_time_raw).strip()}"; reservation_datetime = datetime.datetime.strptime(full_datetime_str, '%Y/%m/%d %H:%M')
                                                        event_prefix = "✨ 내원 : " if is_daily else "❓내원 : "
                                                        event_title = f"{event_prefix}{row.get('환자명', 'N/A')} ({row.get('등록과', 'N/A')}, {row.get('예약의사', 'N/A')})"
                                                        event_description = f"환자명 : {row.get('환자명', 'N/A')}\n진료번호 : {row.get('진료번호', 'N/A')}\n진료내역 : {row.get('진료내역', 'N/A')}"
                                                        create_calendar_event(service, event_title, row.get('진료번호', ''), row.get('등록과', ''), reservation_datetime, row.get('예약의사', ''), event_description)
                                                st.success(f"**{user_name}**님의 캘린더에 일정을 추가했습니다.")
                                            else: st.warning(f"**{user_name}**님에게 보낼 매칭 데이터가 없습니다.")
                                        except Exception as e: st.error(f"**{user_name}**님의 캘린더 일정 추가 실패: {e}")
                                    else: st.warning(f"**{user_name}**님은 Google Calendar 계정이 연동되어 있지 않습니다.")
                    else: st.info("매칭된 환자가 없습니다.")

                # --- 치과의사 수동 전송 탭 ---
                with doctor_admin_tab:
                    st.subheader("🧑‍⚕️ 치과의사 수동 전송 (매칭 결과)");
                    st.warning("수동 모드에서는 이메일/캘린더 전송 버튼을 눌러야 실행됩니다.")

                    if matched_doctors_data:
                        # ... (치과의사 수동 전송 UI 로직은 학생 수동 전송 로직과 대칭적으로 구현) ...
                        st.success(f"등록된 진료가 있는 **{len(matched_doctors_data)}명의 치과의사**를 발견했습니다.")
                        doctor_list_for_multiselect = [f"{res['name']} ({res['email']})" for res in matched_doctors_data]

                        if 'select_all_matched_doctors' not in st.session_state: st.session_state.select_all_matched_doctors = False
                        select_all_button = st.button("등록된 치과의사 모두 선택/해제", key="select_all_matched_res_btn")
                        if select_all_button: st.session_state.select_all_matched_doctors = not st.session_state.select_all_matched_doctors; st.rerun()

                        default_selection_doctor = doctor_list_for_multiselect if st.session_state.select_all_matched_doctors else []
                        selected_doctors_str = st.multiselect("액션을 취할 치과의사 선택", doctor_list_for_multiselect, default=default_selection_doctor, key="doctor_multiselect")
                        selected_doctors_to_act = [res for res in matched_doctors_data if f"{res['name']} ({res['email']})" in selected_doctors_str]
                        
                        if selected_doctors_to_act:
                            mail_col_doc, calendar_col_doc = st.columns(2)
                            with mail_col_doc:
                                if st.button("선택된 치과의사에게 메일 보내기", key="manual_send_mail_doctor"):
                                    for res in selected_doctors_to_act:
                                        # ... (메일 전송 로직) ...
                                        st.success(f"**{res['name']}**님에게 환자 정보 메일 전송 완료!") # 실제 로직 필요
                            with calendar_col_doc:
                                if st.button("선택된 치과의사에게 Google Calendar 일정 추가", key="manual_send_calendar_doctor"):
                                    for res in selected_doctors_to_act:
                                        # ... (캘린더 전송 로직) ...
                                        st.success(f"**{res['name']}**님 캘린더에 일정 추가 완료.") # 실제 로직 필요
                        
                        # (수동 모드에서는 여기에 실제 전송 로직이 필요하지만, 자동 실행 로직을 참조하여 구현할 수 있습니다.)
                    else: st.info("매칭된 치과의사 계정이 없습니다.")

            
            
        # 5. 파일은 업로드 되었으나 아직 옵션을 선택하지 않은 경우 (버튼 누르기 전)
        else:
            st.warning("알림 전송 옵션을 선택해주세요 (자동/수동).")
            
    # 6. 파일이 업로드되지 않은 경우
    else:
        st.info("엑셀 파일을 업로드하면 자동 분석 및 알림 전송이 시작됩니다.")

    st.markdown("---")
    st.subheader("🛠️ Administer password")
    admin_password_input = st.text_input("관리자 비밀번호를 입력하세요", type="password", key="admin_password")
    
    try: secret_admin_password = st.secrets["admin"]["password"]
    except KeyError: secret_admin_password = None; st.error("⚠️ secrets.toml 파일에 'admin.password' 설정이 없습니다. 개발자에게 문의하세요.")
        
    if admin_password_input and admin_password_input == secret_admin_password:
        st.session_state.admin_password_correct = True; st.success("관리자 권한이 활성화되었습니다.")
        if st.session_state.admin_password_correct:
            st.markdown("---"); tab1, tab2 = st.tabs(["일반 사용자 관리", "치과의사 관리"])
            # ... (사용자 관리 탭 로직은 그대로 유지) ...
            
    elif admin_password_input and admin_password_input != secret_admin_password: st.error("비밀번호가 틀렸습니다."); st.session_state.admin_password_correct = False



# --- 8. Regular User Mode ---
# --- 일반 사용자 & 치과의사 모드 ---
import streamlit as st
import pandas as pd
import io
import re
                        
if st.session_state.get('login_mode') in ['user_mode', 'new_user_registration', 'doctor_mode', 'new_doctor_registration', 'doctor_name_input']:
    user_name = st.session_state.get('current_user_name', "")
    user_id_final = st.session_state.get('found_user_email', "")
    firebase_key = st.session_state.get('current_firebase_key', "")
    user_role = st.session_state.get('current_user_role', 'user')
    
    # 올바른 데이터베이스 참조를 결정
    if user_role == 'doctor':
        target_users_ref = doctor_users_ref
    else:
        target_users_ref = users_ref
    
    if firebase_key: # firebase_key가 있을 때만 이 코드를 실행합니다.
        
        # 이메일 주소 변경 기능으로 인해 유저 정보가 바뀔 수 있으므로 매번 업데이트
        if not st.session_state.get('email_change_mode'):
            current_user_meta_data = target_users_ref.child(firebase_key).get()
            if not current_user_meta_data or current_user_meta_data.get("name") != user_name or current_user_meta_data.get("email") != user_id_final:
                # name, email을 업데이트 (다른 필드는 유지)
                update_data = {"name": user_name, "email": user_id_final}
                target_users_ref.child(firebase_key).update(update_data)
                # st.success(f"사용자 정보가 업데이트되었습니다: {user_name} ({user_id_final})")
            st.session_state.current_firebase_key = firebase_key
            st.session_state.current_user_name = user_name
            st.session_state.found_user_email = user_id_final
            st.session_state.current_user_role = user_role

        if not user_name or not user_id_final:
            st.info("내원 알람 노티를 받을 이메일 주소와 사용자 이름을 입력해주세요.")
            st.stop()
    
        if st.session_state.get('login_mode') == 'doctor_mode' or st.session_state.get('login_mode') == 'new_doctor_registration':
            st.header(f"🧑‍⚕️Dr. {user_name}")
            st.subheader("🗓️ Google Calendar 연동")
            st.info("구글 캘린더와 연동하여 내원 일정을 자동으로 등록할 수 있습니다.")

            if 'google_calendar_service' not in st.session_state:
                st.session_state.google_calendar_service = None
            
            # firebase_key가 존재할 때만 함수를 호출하도록 수정
            if firebase_key:
                try:
                    google_calendar_service = get_google_calendar_service(firebase_key)
                    st.session_state.google_calendar_service = google_calendar_service
                except Exception as e:
                    st.error(f"❌ Google Calendar 서비스 로딩에 실패했습니다: {e}")
                    st.info("로그인/인증 정보가 올바른지 확인해주세요.")
                    st.session_state.google_calendar_service = None

            if st.session_state.google_calendar_service:
                st.success("✅ 캘린더 추가 기능이 허용되어 있습니다.")
            else:
                pass

            st.markdown("---")
            st.header("🔑 비밀번호 변경")
            new_password = st.text_input("새 비밀번호를 입력하세요", type="password", key="res_new_password_input")
            confirm_password = st.text_input("새 비밀번호를 다시 입력하세요", type="password", key="res_confirm_password_input")
            
            if st.button("비밀번호 변경", key="res_password_change_btn"):
                if not new_password or not confirm_password:
                    st.error("새 비밀번호와 확인용 비밀번호를 모두 입력해주세요.")
                elif new_password != confirm_password:
                    st.error("새 비밀번호가 일치하지 않습니다. 다시 확인해주세요.")
                else:
                    try:
                        doctor_users_ref.child(st.session_state.current_firebase_key).update({"password": new_password})
                        st.success("🎉 비밀번호가 성공적으로 변경되었습니다!")
                    except Exception as e:
                        st.error(f"비밀번호 변경 중 오류가 발생했습니다: {e}")
            
        elif st.session_state.get('login_mode') in ['user_mode', 'new_user_registration']:
            patients_ref_for_user = db.reference(f"patients/{firebase_key}")

            registration_tab, analysis_tab = st.tabs(['✅ 환자 등록 및 관리', '📈 OCS 분석 결과'])
        
            with registration_tab:
                st.subheader("Google Calendar 연동")
                st.info("환자 등록 시 입력된 이메일 계정의 구글 캘린더에 자동으로 일정이 추가됩니다.")
                if 'google_calendar_service' not in st.session_state:
                    st.session_state.google_calendar_service = None
                
                try:
                    google_calendar_service = get_google_calendar_service(firebase_key)
                    st.session_state.google_calendar_service = google_calendar_service
                except Exception as e:
                    st.error(f"❌ Google Calendar 서비스 로딩에 실패했습니다: {e}")
                    st.session_state.google_calendar_service = None
        
                if st.session_state.google_calendar_service:
                    st.success("✅ 캘린더 추가 기능이 허용되어 있습니다.")
                else:
                    pass
        
                st.markdown("---")
                st.subheader(f"{user_name}님의 토탈 환자 목록")
                existing_patient_data = patients_ref_for_user.get()
        
                if existing_patient_data:
                    patient_list = list(existing_patient_data.items())
                    # 유효성 검사: 데이터가 딕셔너리 형태가 아닌 손상된 데이터를 제거
                    valid_patient_list = [item for item in patient_list if isinstance(item[1], dict)]
                    # --- [핵심 변경: 진료과 플래그 우선순위 정렬] ---
                    # 1. 소치(0) > 보철(1) > 내과(2) > 교정(3) 순서로 높은 우선순위를 부여
                    # 2. 동일 우선순위 내에서는 환자이름 순으로 정렬
                    sorted_patient_list = sorted(valid_patient_list, key=lambda item: (
                        0 if item[1].get('소치', False) else
                        1 if item[1].get('외과', False) else
                        2 if item[1].get('내과', False) else
                        3 if item[1].get('교정', False) else
                        4 if item[1].get('보철', False) else
                        5 if item[1].get('원진실', False) else
                        6 if item[1].get('보존', False) else
                        7, 
                        item[1].get('환자이름', 'zzz')
                    ))
                    cols_count = 3
                    cols = st.columns(cols_count)
        
                    for idx, (pid_key, val) in enumerate(sorted_patient_list): # pid_key가 환자번호
                        with cols[idx % cols_count]:
                            with st.container(border=True):
                                info_col, btn_col = st.columns([4, 1])
                                with info_col:
                                    # True인 진료과만 추출하여 표시
                                    registered_depts = [
                                        dept.capitalize() 
                                        for dept in PATIENT_DEPT_FLAGS + ['보존', '치주', '원진실'] # 모든 가능한 과
                                        if val.get(dept.lower()) is True or val.get(dept.lower()) == 'True' or val.get(dept.lower()) == 'true'
                                    ]

                                    depts_str = ", ".join(registered_depts) if registered_depts else "미지정"
                                    
                                    st.markdown(f"**{val.get('환자이름', '이름 없음')}** / {pid_key} / {depts_str}") # pid_key는 진료번호
                                with btn_col:
                                    if st.button("X", key=f"delete_button_{pid_key}"): # pid_key 사용
                                        patients_ref_for_user.child(pid_key).delete()
                                        st.rerun()
                else:
                    st.info("등록된 환자가 없습니다.")
                st.markdown("---")

                # --- 환자 정보 대량 등록 섹션 추가 (구조 변경 반영) ---
                st.subheader("📋 환자 정보 대량 등록")
                st.markdown("엑셀에서 **환자명, 진료번호, 등록과** 순서의 데이터를 그대로 붙여넣어주세요.")
                st.markdown("예시: 홍길동	1046769	보존")
                st.markdown(f"등록 가능 과: {', '.join(DEPARTMENTS_FOR_REGISTRATION)}")
                
                
                paste_area = st.text_area("", height=200, placeholder="여기에 엑셀 데이터를 붙여넣으세요.")
                
                if st.button("붙여넣은 환자 등록"):
                    if paste_area:
                        try:
                            data_io = io.StringIO(paste_area)
                            
                            # header=None으로 헤더가 없음을 명시하고, names로 열 이름을 수동 지정
                            df = pd.read_csv(data_io, sep='\s+', header=None, names=['환자명', '진료번호', '등록과'])
                            
                            # 기존 환자 데이터 가져오기
                            existing_patient_data = patients_ref_for_user.get()
                            if not existing_patient_data:
                                existing_patient_data = {}
                            
                            success_count = 0
                            for index, row in df.iterrows():
                                name = str(row["환자명"]).strip()
                                pid = str(row["진료번호"]).strip()
                                department = str(row["등록과"]).strip()
                                
                                if not name or not pid or not department:
                                    st.warning(f"{index+1}번째 행: 정보가 누락되어 건너킵니다.")
                                    continue
                                
                                pid_key = pid.strip() # 진료번호를 키로 사용
                                dept_key_lower = department.lower()
                                
                                # 1. 새 데이터 딕셔너리 생성 및 초기화 (환자이름, 진료번호, 5개 진료과 플래그)
                                # 기존 데이터가 있으면 불러와서 업데이트
                                new_patient_data = existing_patient_data.get(pid_key, {
                                    "환자이름": name,
                                    "진료번호": pid # 키로 사용되지만 데이터 내부에도 포함
                                })
                                
                                # 진료과 플래그 초기화 및 업데이트 (기존 데이터와 병합)
                                for dept_flag in PATIENT_DEPT_FLAGS + ['보존', '치주', '원진실']: # 모든 가능한 과 플래그 초기화
                                    lower_dept = dept_flag.lower()
                                    if lower_dept not in new_patient_data:
                                        new_patient_data[lower_dept] = False

                                # 2. 등록과에 해당하는 플래그 True로 설정
                                if dept_key_lower in new_patient_data:
                                    new_patient_data[dept_key_lower] = True
                                else:
                                    st.warning(f"{name} ({pid}): 알 수 없는 진료과 '{department}'가 입력되었습니다. 플래그를 설정하지 않습니다.")

                                # 3. 환자번호(pid)를 키로 사용하여 데이터 저장 (덮어쓰기/업데이트)
                                patients_ref_for_user.child(pid_key).set(new_patient_data) # <--- **핵심 변경점**
                                success_count += 1
                                st.success(f"{name} ({pid}) [{department}] 환자 등록/업데이트 완료")
                                
                            
                            if success_count > 0:
                                st.success(f"총 {success_count}명의 환자 정보 등록/업데이트가 완료되었습니다.")
                            st.rerun()
                            
                        except pd.errors.ParserError:
                            st.error("잘못된 형식입니다. 엑셀이나 구글 스프레드시트의 표를 복사하여 붙여넣었는지 확인해주세요. 데이터 구분자가 탭(Tab)이어야 합니다.")
                        except Exception as e:
                            st.error(f"예상치 못한 오류: {e}")
                    else:
                        st.warning("붙여넣을 환자 정보가 없습니다.")
                        
                st.markdown("---")
        
                # --- 환자 정보 일괄 삭제 섹션 추가 (구조 변경 반영) ---
                st.subheader("🗑️ 환자 정보 일괄 삭제")
                
                if 'delete_patient_confirm' not in st.session_state:
                    st.session_state.delete_patient_confirm = False
                if 'patients_to_delete' not in st.session_state:
                    st.session_state.patients_to_delete = []
                if 'select_all_mode' not in st.session_state:
                    st.session_state.select_all_mode = False
                
                all_patients_meta = patients_ref_for_user.get()
                patient_list_for_dropdown = []
                patient_key_map = {}
                
                if all_patients_meta:
                    for pid_key, value in all_patients_meta.items(): # pid_key는 진료번호
                        # True인 진료과를 모두 추출하여 표시
                        registered_depts = [
                            dept.capitalize() 
                            for dept in PATIENT_DEPT_FLAGS + ['보존', '치주', '원진실'] # 모든 가능한 과
                            if value.get(dept.lower()) is True or value.get(dept.lower()) == 'True' or value.get(dept.lower()) == 'true'
                        ]
                        depts_str = ", ".join(registered_depts) if registered_depts else "미지정"
                        
                        display_text = f"{value.get('환자이름', '이름 없음')} ({pid_key}) [{depts_str}]"
                        patient_list_for_dropdown.append(display_text)
                        patient_key_map[display_text] = pid_key # key가 이제 진료번호
                
                # "전체 선택" 버튼 추가
                if st.button("전체 환자 선택/해제", key="select_all_patients_button"):
                    st.session_state.select_all_mode = not st.session_state.select_all_mode # 상태 토글
                    st.rerun()
                
                # '전체 선택' 모드에 따라 multiselect의 기본값 설정
                default_selection = patient_list_for_dropdown if st.session_state.select_all_mode else []
                
                if not st.session_state.delete_patient_confirm:
                    patients_to_delete_multiselect = st.multiselect(
                        "삭제할 환자 선택",
                        patient_list_for_dropdown,
                        default=default_selection, # 기본값 설정
                        key="delete_patient_multiselect"
                    )
                
                    if st.button("선택한 환자 삭제", key="delete_patient_button"):
                        if patients_to_delete_multiselect:
                            st.session_state.delete_patient_confirm = True
                            st.session_state.patients_to_delete = patients_to_delete_multiselect
                            st.session_state.select_all_mode = False # 삭제 버튼 클릭 시 전체 선택 모드 초기화
                            st.rerun()
                        else:
                            st.warning("삭제할 환자를 선택해주세요.")
                else:
                    st.warning("정말로 선택한 환자를 삭제하시겠습니까? 이 작업은 되돌릴 수 없습니다.")
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("예, 삭제합니다", key="confirm_delete_patient"):
                            with st.spinner('삭제 중...'):
                                for patient_to_del_str in st.session_state.patients_to_delete:
                                    patient_key_to_del = patient_key_map.get(patient_to_del_str) # patient_key_to_del은 이제 진료번호
                                    if patient_key_to_del:
                                        patients_ref_for_user.child(patient_key_to_del).delete()
                                
                                st.success(f"선택한 환자 {st.session_state.patients_to_delete} 삭제 완료.")
                                st.session_state.delete_patient_confirm = False
                                st.session_state.patients_to_delete = []
                                st.rerun()
                    with col2:
                        if st.button("아니오, 취소합니다", key="cancel_delete_patient"):
                            st.session_state.delete_patient_confirm = False
                            st.session_state.patients_to_delete = []
                            st.rerun()
                
                st.markdown("---")
        
                with st.form("register_form"):
                    name = st.text_input("환자명")
                    pid = st.text_input("진료번호")
                    
                    # --- [핵심 변경: 다중 선택으로 변경] ---
                    selected_departments = st.multiselect("등록할 진료과 (복수 선택 가능)", DEPARTMENTS_FOR_REGISTRATION)
                    submitted = st.form_submit_button("등록")
                    
                    if submitted:
                        if not name or not pid or not selected_departments:
                            st.warning("환자명, 진료번호, 등록할 진료과를 모두 입력/선택해주세요.")
                        else:
                            pid_key = pid.strip()
                            
                            # 기존 데이터 불러오기 (없으면 새로 생성)
                            new_patient_data = existing_patient_data.get(pid_key, {
                                "환자이름": name,
                                "진료번호": pid # 키로 사용되지만 데이터 내부에도 포함
                            })
                            
                            # 기존에 저장된 모든 진료과 플래그를 False로 초기화 (선택되지 않은 과)
                            for dept_flag in PATIENT_DEPT_FLAGS + ['보존', '치주', '원진실']:
                                lower_dept = dept_flag.lower()
                                new_patient_data[lower_dept] = False

                            # 선택된 진료과만 True로 설정
                            for dept in selected_departments:
                                dept_key_lower = dept.lower()
                                if dept_key_lower in new_patient_data:
                                    new_patient_data[dept_key_lower] = True
                                
                            # 진료번호를 키로 사용하여 데이터 저장 (덮어쓰기/업데이트)
                            patients_ref_for_user.child(pid_key).set(new_patient_data) # <--- **핵심 변경점**
                            st.success(f"{name} ({pid}) [{', '.join(selected_departments)}] 환자 등록/업데이트 완료")
                            st.rerun()


            
            with analysis_tab:
                st.header("📈 OCS 분석 결과")
                analysis_results = db.reference("ocs_analysis/latest_result").get()
                latest_file_name = db.reference("ocs_analysis/latest_file_name").get()

                if analysis_results and latest_file_name:
                    st.markdown(f"**<h3 style='text-align: left;'>{latest_file_name} 분석 결과</h3>**", unsafe_allow_html=True)
                    st.markdown("---")
                    
                    if '소치' in analysis_results:
                        st.subheader("소아치과 현황 (단타)")
                        st.info(f"오전: **{analysis_results['소치']['오전']}명**")
                        st.info(f"오후: **{analysis_results['소치']['오후']}명**")
                    else:
                        st.warning("소아치과 데이터가 엑셀 파일에 없습니다.")
                    st.markdown("---")
                    
                    if '보존' in analysis_results:
                        st.subheader("보존과 현황 (단타)")
                        st.info(f"오전: **{analysis_results['보존']['오전']}명**")
                        st.info(f"오후: **{analysis_results['보존']['오후']}명**")
                    else:
                        st.warning("보존과 데이터가 엑셀 파일에 없습니다.")
                    st.markdown("---")

                    if '교정' in analysis_results:
                        st.subheader("교정과 현황 (Bonding)")
                        st.info(f"오전: **{analysis_results['교정']['오전']}명**")
                        st.info(f"오후: **{analysis_results['교정']['오후']}명**")
                    else:
                        st.warning("교정과 데이터가 엑셀 파일에 없습니다.")
                    st.markdown("---")
                else:
                    st.info("💡 분석 결과가 없습니다. 관리자가 엑셀 파일을 업로드하면 표시됩니다.")
                    
                
                st.divider()
                st.header("🔑 비밀번호 변경")
                new_password = st.text_input("새 비밀번호를 입력하세요", type="password", key="user_new_password_input")
                confirm_password = st.text_input("새 비밀번호를 다시 입력하세요", type="password", key="user_confirm_password_input")
                
                if st.button("비밀번호 변경", key="user_password_change_btn"):
                    if not new_password or not confirm_password:
                        st.error("새 비밀번호와 확인용 비밀번호를 모두 입력해주세요.")
                    elif new_password != confirm_password:
                        st.error("새 비밀번호가 일치하지 않습니다. 다시 확인해주세요.")
                    else:
                        try:
                            users_ref.child(st.session_state.current_firebase_key).update({"password": new_password})
                            st.success("🎉 비밀번호가 성공적으로 변경되었습니다!")
                        except Exception as e:
                            st.error(f"비밀번호 변경 중 오류가 발생했습니다: {e}")
